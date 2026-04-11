import tkinter as tk
from datetime import date, datetime, timedelta
from tkinter import filedialog, messagebox, ttk
from typing import Callable

from config.config import (
    COLORS,
    CONDICIONES_FILE,
    ENTRADAS_FILE,
    PROVEEDORES_FILE,
    REPORTES_PATH,
    SALIDAS_FILE,
    SUSTANCIAS_FILE,
    TIPOS_ENTRADA_FILE,
    TIPOS_SALIDA_FILE,
    UBICACIONES_FILE,
    UBICACIONES_USO_FILE,
    UNIDADES_FILE,
)
from ui.styles import build_header
from utils.data_handler import (
    DataHandler,
    Lookups,
    build_location_indexes,
    build_substance_indexes,
    location_name,
    substance_code,
    substance_code_system,
    substance_name,
)

try:
    from openpyxl import Workbook, load_workbook
except ImportError:
    Workbook = None
    load_workbook = None


MESES = [
    "Enero",
    "Febrero",
    "Marzo",
    "Abril",
    "Mayo",
    "Junio",
    "Julio",
    "Agosto",
    "Septiembre",
    "Octubre",
    "Noviembre",
    "Diciembre",
]


class ReportesWindow:
    """Generación de reportes por mes, rango o histórico en Excel."""

    def __init__(self, parent: tk.Tk):
        self.window = tk.Toplevel(parent)
        self.window.title("Reportes")
        self.window.geometry("760x340")
        self.window.configure(bg=COLORS["secondary"])

        today = date.today()
        self.mes_var = tk.StringVar(value=MESES[today.month - 1])
        self.anio_var = tk.StringVar(value=str(today.year))
        self.filtro_var = tk.StringVar(value="mes")
        self.desde_var = tk.StringVar(value=today.strftime("%Y-%m-01"))
        self.hasta_var = tk.StringVar(value=today.strftime("%Y-%m-%d"))

        self._mes_combo: ttk.Combobox | None = None
        self._anio_combo: ttk.Combobox | None = None
        self._desde_entry: tk.Entry | None = None
        self._hasta_entry: tk.Entry | None = None

        self._build_ui()

    def _build_ui(self) -> None:
        wrapper = tk.Frame(self.window, bg="white", bd=1, relief="solid", padx=12, pady=12)
        wrapper.pack(expand=True, fill="both", padx=14, pady=14)

        build_header(wrapper, "Sistema de Gestión  -  Reportes")

        # ── Fecha ──
        box_fecha = tk.LabelFrame(wrapper, text="Período", bg="white", fg="#1F4F8A", font=("Segoe UI", 10, "bold"))
        box_fecha.pack(fill="x", pady=(8, 8), padx=4)

        row = tk.Frame(box_fecha, bg="white")
        row.pack(fill="x", padx=10, pady=10)

        tk.Radiobutton(
            row,
            text="Mes/Año",
            variable=self.filtro_var,
            value="mes",
            bg="white",
            command=self._update_filter_state,
        ).grid(row=0, column=0, sticky="w", padx=(0, 10))

        tk.Radiobutton(
            row,
            text="Rango",
            variable=self.filtro_var,
            value="rango",
            bg="white",
            command=self._update_filter_state,
        ).grid(row=0, column=1, sticky="w", padx=(0, 10))

        tk.Radiobutton(
            row,
            text="Todo histórico",
            variable=self.filtro_var,
            value="todo",
            bg="white",
            command=self._update_filter_state,
        ).grid(row=0, column=2, sticky="w")

        row2 = tk.Frame(box_fecha, bg="white")
        row2.pack(fill="x", padx=10, pady=(0, 10))

        tk.Label(row2, text="Mes", bg="white").grid(row=0, column=0, padx=(0, 6), sticky="w")
        self._mes_combo = ttk.Combobox(row2, textvariable=self.mes_var, values=MESES, state="readonly", width=16)
        self._mes_combo.grid(row=0, column=1, padx=(0, 14), sticky="w")

        tk.Label(row2, text="Año", bg="white").grid(row=0, column=2, padx=(0, 6), sticky="w")
        self._anio_combo = ttk.Combobox(row2, textvariable=self.anio_var, values=self._available_years(), state="readonly", width=10)
        self._anio_combo.grid(row=0, column=3, sticky="w", padx=(0, 20))

        tk.Label(row2, text="Desde", bg="white").grid(row=0, column=4, padx=(0, 6), sticky="w")
        self._desde_entry = tk.Entry(row2, textvariable=self.desde_var, width=12)
        self._desde_entry.grid(row=0, column=5, padx=(0, 10), sticky="w")

        tk.Label(row2, text="Hasta", bg="white").grid(row=0, column=6, padx=(0, 6), sticky="w")
        self._hasta_entry = tk.Entry(row2, textvariable=self.hasta_var, width=12)
        self._hasta_entry.grid(row=0, column=7, sticky="w")

        tk.Label(
            box_fecha,
            text="Formato rango: YYYY-MM-DD",
            bg="white",
            fg="#666666",
            font=("Segoe UI", 9),
        ).pack(anchor="w", padx=10, pady=(0, 8))
        self._update_filter_state()

        actions = tk.Frame(wrapper, bg="white")
        actions.pack(fill="x", pady=(8, 0))

        tk.Button(
            actions,
            text="Consolidado",
            command=self.generate_report,
            bg=COLORS["primary"],
            fg=COLORS["text_light"],
            relief="flat",
            padx=20,
            pady=7,
        ).pack(side="left", padx=(0, 8))

        tk.Button(
            actions,
            text="Reporte Entradas",
            command=self.generate_entradas_report,
            bg=COLORS["primary"],
            fg=COLORS["text_light"],
            relief="flat",
            padx=20,
            pady=7,
        ).pack(side="left", padx=(0, 8))

        tk.Button(
            actions,
            text="Reporte Salidas",
            command=self.generate_salidas_report,
            bg=COLORS["primary"],
            fg=COLORS["text_light"],
            relief="flat",
            padx=20,
            pady=7,
        ).pack(side="left")

        tk.Button(
            actions,
            text="Salir",
            command=self.window.destroy,
            bg=COLORS["border"],
            fg=COLORS["text_dark"],
            relief="flat",
            padx=20,
            pady=7,
        ).pack(side="right")

    def _update_filter_state(self) -> None:
        mode = self.filtro_var.get()
        month_state = "readonly" if mode == "mes" else "disabled"
        range_state = "normal" if mode == "rango" else "disabled"

        if self._mes_combo is not None:
            self._mes_combo.configure(state=month_state)
        if self._anio_combo is not None:
            self._anio_combo.configure(state=month_state)
        if self._desde_entry is not None:
            self._desde_entry.configure(state=range_state)
        if self._hasta_entry is not None:
            self._hasta_entry.configure(state=range_state)

    def _resolve_filter(self, date_field: str) -> tuple[str, date, str, Callable[[dict], bool]] | None:
        mode = self.filtro_var.get().strip()

        if mode == "mes":
            period = self._validated_period()
            if period is None:
                return None
            mes_nombre, anio, mes = period
            corte = _month_end(anio, mes)
            suffix = f"{anio}_{mes:02d}"

            def matcher(rec: dict) -> bool:
                return _is_in_month(rec.get(date_field, ""), anio, mes)

            return mes_nombre, corte, suffix, matcher

        if mode == "rango":
            desde = _parse_iso_date(self.desde_var.get())
            hasta = _parse_iso_date(self.hasta_var.get())
            if desde is None or hasta is None:
                messagebox.showerror(
                    "Validación",
                    "Rango inválido. Usa formato YYYY-MM-DD en Desde y Hasta.",
                    parent=self.window,
                )
                return None
            if desde > hasta:
                messagebox.showerror(
                    "Validación",
                    "La fecha Desde no puede ser mayor que Hasta.",
                    parent=self.window,
                )
                return None

            periodo_label = f"{desde:%Y-%m-%d} a {hasta:%Y-%m-%d}"
            suffix = f"{desde:%Y%m%d}_{hasta:%Y%m%d}"

            def matcher(rec: dict) -> bool:
                rec_date = _parse_iso_date(rec.get(date_field, ""))
                return rec_date is not None and desde <= rec_date <= hasta

            return periodo_label, hasta, suffix, matcher

        if mode == "todo":
            def matcher(_rec: dict) -> bool:
                return True

            return "Histórico", date.today(), "historico", matcher

        messagebox.showerror("Validación", "Modo de filtro no válido", parent=self.window)
        return None

    def _validated_period(self) -> tuple[str, int, int] | None:
        mes_nombre = self.mes_var.get().strip()
        anio_raw = self.anio_var.get().strip()

        if mes_nombre not in MESES:
            messagebox.showerror("Validación", "Selecciona un mes válido", parent=self.window)
            return None

        try:
            anio = int(anio_raw)
        except ValueError:
            messagebox.showerror("Validación", "Selecciona un año válido", parent=self.window)
            return None

        mes = MESES.index(mes_nombre) + 1
        return mes_nombre, anio, mes

    def _available_years(self) -> list[str]:
        years = {str(date.today().year)}
        for rec in DataHandler.get_all(ENTRADAS_FILE, "entradas"):
            y = _extract_year(rec.get("fecha", ""))
            if y:
                years.add(str(y))
        for rec in DataHandler.get_all(SALIDAS_FILE, "salidas"):
            y = _extract_year(rec.get("fecha_salida", ""))
            if y:
                years.add(str(y))
        return sorted(years)

    def generate_report(self) -> None:
        if Workbook is None or load_workbook is None:
            messagebox.showerror(
                "Dependencia faltante",
                "No se encontró openpyxl. Instala con: pip install openpyxl",
                parent=self.window,
            )
            return

        resolved_entradas = self._resolve_filter("fecha")
        resolved_salidas = self._resolve_filter("fecha_salida")
        if resolved_entradas is None or resolved_salidas is None:
            return

        periodo_label, corte, suffix, entradas_match = resolved_entradas
        _, _, _, salidas_match = resolved_salidas

        # Ask where to save
        default_name = f"reporte_consolidado_{suffix}.xlsx"
        save_path = filedialog.asksaveasfilename(
            title="Guardar reporte",
            initialfile=default_name,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialdir=str(REPORTES_PATH),
            parent=self.window,
        )
        if not save_path:
            return

        sustancias = DataHandler.load_json(SUSTANCIAS_FILE).get("maestrasSustancias", [])
        unidades = DataHandler.load_json(UNIDADES_FILE).get("maestrasUnidades", [])
        tipos_entrada = DataHandler.load_json(TIPOS_ENTRADA_FILE).get("maestrasTiposEntrada", [])
        tipos_salida = DataHandler.load_json(TIPOS_SALIDA_FILE).get("maestrasTiposSalida", [])

        sustancias_by_id = {int(x.get("id")): x for x in sustancias if x.get("id") is not None}
        unidades_by_id = {int(x.get("id")): x for x in unidades if x.get("id") is not None}
        tipos_entrada_by_id = {int(x.get("id")): str(x.get("nombre", "")) for x in tipos_entrada if x.get("id") is not None}
        tipos_salida_by_id = {int(x.get("id")): str(x.get("nombre", "")) for x in tipos_salida if x.get("id") is not None}

        entradas = [r for r in DataHandler.get_all(ENTRADAS_FILE, "entradas") if entradas_match(r)]
        salidas = [r for r in DataHandler.get_all(SALIDAS_FILE, "salidas") if salidas_match(r)]

        entradas_rows = _build_entradas_rows(entradas, sustancias_by_id, unidades_by_id, tipos_entrada_by_id)
        salidas_rows = _build_salidas_rows(salidas, sustancias_by_id, unidades_by_id, tipos_salida_by_id)

        wb = Workbook()
        wb.remove(wb.active)
        
        template_path = REPORTES_PATH / "templates" / "template_reporte.xlsx"

        _render_combined_sheet(
            wb,
            "Reporte",
            template_path,
            corte,
            periodo_label,
            entradas_rows,
            salidas_rows,
        )

        try:
            wb.save(save_path)
            messagebox.showinfo(
                "Reporte generado",
                f"Excel generado correctamente:\n{save_path}",
                parent=self.window,
            )
        except Exception as e:
            messagebox.showerror(
                "Error al guardar",
                f"No se pudo guardar el archivo:\n{str(e)}",
                parent=self.window,
            )

    def generate_entradas_report(self) -> None:
        if load_workbook is None:
            messagebox.showerror(
                "Dependencia faltante",
                "No se encontró openpyxl. Instala con: pip install openpyxl",
                parent=self.window,
            )
            return

        resolved = self._resolve_filter("fecha")
        if resolved is None:
            return
        _, _, suffix, matcher = resolved

        template_path = REPORTES_PATH / "templates" / "template_reporteentrada.xlsx"
        if not template_path.exists():
            messagebox.showerror(
                "Plantilla no encontrada",
                f"No existe la plantilla:\n{template_path}",
                parent=self.window,
            )
            return

        default_name = f"reporte_entradas_{suffix}.xlsx"
        save_path = filedialog.asksaveasfilename(
            title="Guardar reporte de entradas",
            initialfile=default_name,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialdir=str(REPORTES_PATH),
            parent=self.window,
        )
        if not save_path:
            return

        entradas = [
            r for r in DataHandler.get_all(ENTRADAS_FILE, "entradas")
            if not r.get("anulado") and matcher(r)
        ]

        unidades_cat = DataHandler.load_json(UNIDADES_FILE).get("maestrasUnidades", [])
        proveedores_cat = DataHandler.load_json(PROVEEDORES_FILE).get("maestrasProveedores", [])
        condiciones_cat = DataHandler.load_json(CONDICIONES_FILE).get("maestrasCondicionesAlmacenamiento", [])
        sustancias_cat = DataHandler.load_json(SUSTANCIAS_FILE).get("maestrasSustancias", [])
        ubicaciones_cat = DataHandler.load_json(UBICACIONES_FILE).get("maestrasUbicaciones", [])
        ubicaciones_uso_cat = DataHandler.load_json(UBICACIONES_USO_FILE).get("maestrasUbicacionesUso", [])
        tipos_entrada_cat = DataHandler.load_json(TIPOS_ENTRADA_FILE).get("maestrasTiposEntrada", [])

        lkp = Lookups(
            unidades=unidades_cat,
            proveedores=proveedores_cat,
            condiciones=condiciones_cat,
            tipos_entrada=tipos_entrada_cat,
        )
        sustancias_by_id, _ = build_substance_indexes(sustancias_cat)
        locations_by_key, _ = build_location_indexes(ubicaciones_cat, ubicaciones_uso_cat)

        rows = _build_entradas_detailed_rows(entradas, lkp, sustancias_by_id, locations_by_key)

        try:
            _render_template_table(template_path, save_path, rows)
            messagebox.showinfo(
                "Reporte generado",
                f"Excel generado correctamente:\n{save_path}",
                parent=self.window,
            )
        except Exception as e:
            messagebox.showerror(
                "Error al guardar",
                f"No se pudo generar el reporte:\n{e}",
                parent=self.window,
            )

    def generate_salidas_report(self) -> None:
        if load_workbook is None:
            messagebox.showerror(
                "Dependencia faltante",
                "No se encontró openpyxl. Instala con: pip install openpyxl",
                parent=self.window,
            )
            return

        resolved = self._resolve_filter("fecha_salida")
        if resolved is None:
            return
        _, _, suffix, matcher = resolved

        template_path = REPORTES_PATH / "templates" / "template_reportesalida.xlsx"
        if not template_path.exists():
            messagebox.showerror(
                "Plantilla no encontrada",
                f"No existe la plantilla:\n{template_path}",
                parent=self.window,
            )
            return

        default_name = f"reporte_salidas_{suffix}.xlsx"
        save_path = filedialog.asksaveasfilename(
            title="Guardar reporte de salidas",
            initialfile=default_name,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialdir=str(REPORTES_PATH),
            parent=self.window,
        )
        if not save_path:
            return

        entradas_all = [r for r in DataHandler.get_all(ENTRADAS_FILE, "entradas") if not r.get("anulado")]
        salidas_all = [r for r in DataHandler.get_all(SALIDAS_FILE, "salidas") if not r.get("anulado")]
        salidas = [r for r in salidas_all if matcher(r)]

        unidades_cat = DataHandler.load_json(UNIDADES_FILE).get("maestrasUnidades", [])
        sustancias_cat = DataHandler.load_json(SUSTANCIAS_FILE).get("maestrasSustancias", [])
        tipos_salida_cat = DataHandler.load_json(TIPOS_SALIDA_FILE).get("maestrasTiposSalida", [])
        ubicaciones_cat = DataHandler.load_json(UBICACIONES_FILE).get("maestrasUbicaciones", [])
        ubicaciones_uso_cat = DataHandler.load_json(UBICACIONES_USO_FILE).get("maestrasUbicacionesUso", [])

        lkp = Lookups(unidades=unidades_cat, tipos_salida=tipos_salida_cat)
        sustancias_by_id, _ = build_substance_indexes(sustancias_cat)
        locations_by_key, _ = build_location_indexes(ubicaciones_cat, ubicaciones_uso_cat)

        rows = _build_salidas_detailed_rows(
            salidas,
            entradas_all,
            salidas_all,
            lkp,
            sustancias_by_id,
            locations_by_key,
        )

        try:
            _render_template_table(template_path, save_path, rows)
            messagebox.showinfo(
                "Reporte generado",
                f"Excel generado correctamente:\n{save_path}",
                parent=self.window,
            )
        except Exception as e:
            messagebox.showerror(
                "Error al guardar",
                f"No se pudo generar el reporte:\n{e}",
                parent=self.window,
            )


HEADER_COLUMNS = [
    "CÓDIGO DE USO",
    "Código Sistema Contable",
    "NOMBRE",
    "LOTE",
    "CANTIDAD",
    "Unidades",
    "Actividad",
    "SC / NC",
    "Observación",
]


def _build_entradas_rows(
    entradas: list[dict],
    sustancias_by_id: dict[int, dict],
    unidades_by_id: dict[int, dict],
    tipos_entrada_by_id: dict[int, str],
) -> list[list]:
    rows: list[list] = []
    for rec in entradas:
        sust = sustancias_by_id.get(int(rec.get("id_sustancia", 0) or 0), {})
        uni = unidades_by_id.get(int(rec.get("id_unidad", 0) or 0), {})
        cantidad_reporte = rec.get("total", rec.get("cantidad", 0))
        rows.append(
            [
                sust.get("codigo", ""),
                sust.get("codigo_sistema", ""),
                sust.get("nombre", ""),
                rec.get("lote", ""),
                cantidad_reporte,
                uni.get("codigo", ""),
                tipos_entrada_by_id.get(int(rec.get("id_tipo_entrada", 0) or 0), ""),
                rec.get("sc_nc", "No controlada"),
                rec.get("observaciones", "") or "N/A",
            ]
        )
    return rows


def _build_salidas_rows(
    salidas: list[dict],
    sustancias_by_id: dict[int, dict],
    unidades_by_id: dict[int, dict],
    tipos_salida_by_id: dict[int, str],
) -> list[list]:
    rows: list[list] = []
    for rec in salidas:
        sust = sustancias_by_id.get(int(rec.get("id_sustancia", 0) or 0), {})
        uni = unidades_by_id.get(int(rec.get("id_unidad", 0) or 0), {})
        rows.append(
            [
                sust.get("codigo", ""),
                sust.get("codigo_sistema", ""),
                sust.get("nombre", ""),
                rec.get("lote", ""),
                rec.get("cantidad", 0),
                uni.get("codigo", ""),
                tipos_salida_by_id.get(int(rec.get("id_tipo_salida", 0) or 0), ""),
                rec.get("sc_nc", "No controlada"),
                rec.get("observaciones", "") or "N/A",
            ]
        )
    return rows


def _render_combined_sheet(
    wb,
    sheet_name: str,
    template_path,
    fecha_reporte: date,
    mes_nombre: str,
    entradas_rows: list[list],
    salidas_rows: list[list],
):
    """Render one sheet from template: entradas section + dynamic salidas section below."""

    try:
        template_wb = load_workbook(template_path)
        template_ws = template_wb.active

        ws = wb.create_sheet(sheet_name)

        # Copy template cells and styles.
        for row in template_ws.iter_rows():
            for cell in row:
                new_cell = ws.cell(row=cell.row, column=cell.column, value=cell.value)
                if cell.has_style:
                    new_cell.font = cell.font.copy() if cell.font else None
                    new_cell.border = cell.border.copy() if cell.border else None
                    new_cell.fill = cell.fill.copy() if cell.fill else None
                    new_cell.number_format = cell.number_format
                    new_cell.alignment = cell.alignment.copy() if cell.alignment else None
        # Copy images (logo)
        from openpyxl.drawing.image import Image

        if hasattr(template_ws, "_images"):
            for img in template_ws._images:
                try:
                    new_img = Image(img.ref)

                    # Mantener posición original
                    new_img.anchor = img.anchor

                    ws.add_image(new_img)
                except Exception:
                    pass
        # Copy merged cells and dimensions from template.
        for merged in template_ws.merged_cells.ranges:
            ws.merge_cells(str(merged))

        for col in template_ws.column_dimensions:
            ws.column_dimensions[col].width = template_ws.column_dimensions[col].width

        for row_dim in template_ws.row_dimensions:
            ws.row_dimensions[row_dim].height = template_ws.row_dimensions[row_dim].height

        ENTRADAS_START_ROW = 11
        SALIDAS_START_ROW = 33

        # Fill general data in template placeholders.
        ws["C7"] = fecha_reporte.strftime("%Y-%m-%d")
        ws["G7"] = mes_nombre
        ws["C29"] = fecha_reporte.strftime("%Y-%m-%d")
        ws["G29"] = mes_nombre

        # ENTRADAS section (fixed block from template).
        start_entradas = ENTRADAS_START_ROW
        for i, row_data in enumerate(entradas_rows):
            target_row = start_entradas + i
            for j, value in enumerate(row_data, start=1):
                base_cell = template_ws.cell(row=ENTRADAS_START_ROW, column=j)
                target = ws.cell(row=target_row, column=j, value=value)
                if base_cell.has_style:
                    target.font = base_cell.font.copy() if base_cell.font else None
                    target.border = base_cell.border.copy() if base_cell.border else None
                    target.fill = base_cell.fill.copy() if base_cell.fill else None
                    target.number_format = base_cell.number_format
                    target.alignment = base_cell.alignment.copy() if base_cell.alignment else None

        # SALIDAS section (fixed block from template).
        start_salidas = SALIDAS_START_ROW
        for i, row_data in enumerate(salidas_rows):
            target_row = start_salidas + i
            for j, value in enumerate(row_data, start=1):
                source = template_ws.cell(row=SALIDAS_START_ROW, column=j)
                target = ws.cell(row=target_row, column=j, value=value)
                if source.has_style:
                    target.font = source.font.copy() if source.font else None
                    target.border = source.border.copy() if source.border else None
                    target.fill = source.fill.copy() if source.fill else None
                    target.number_format = source.number_format
                    target.alignment = source.alignment.copy() if source.alignment else None

        return ws
    except Exception as e:
        import traceback
        traceback.print_exc()
        raise


def _render_template_table(template_path, save_path: str, data_rows: list[list], start_row: int = 2) -> None:
    wb = load_workbook(template_path)
    ws = wb.active

    max_cols = ws.max_column
    style_templates = []
    for col in range(1, max_cols + 1):
        source = ws.cell(row=start_row, column=col)
        style_templates.append(
            {
                "font": source.font.copy() if source.font else None,
                "border": source.border.copy() if source.border else None,
                "fill": source.fill.copy() if source.fill else None,
                "number_format": source.number_format,
                "alignment": source.alignment.copy() if source.alignment else None,
            }
        )

    for i, row_data in enumerate(data_rows):
        current_row = start_row + i
        ws.insert_rows(current_row)
        normalized = list(row_data[:max_cols])
        if len(normalized) < max_cols:
            normalized.extend([""] * (max_cols - len(normalized)))

        for col, value in enumerate(normalized, start=1):
            target = ws.cell(row=current_row, column=col, value=value)
            st = style_templates[col - 1]
            target.font = st["font"]
            target.border = st["border"]
            target.fill = st["fill"]
            target.number_format = st["number_format"]
            target.alignment = st["alignment"]

    if data_rows:
        ws.delete_rows(start_row + len(data_rows))

    wb.save(save_path)


def _bool_label(value) -> str:
    return "Sí" if bool(value) else "No"


def _build_entradas_detailed_rows(
    entradas: list[dict],
    lkp: Lookups,
    sustancias_by_id: dict[int, dict],
    locations_by_key: dict[tuple[str, int], dict],
) -> list[list]:
    rows: list[list] = []
    for rec in entradas:
        rows.append(
            [
                lkp.to_name("tipos_entrada", rec.get("id_tipo_entrada"), ""),
                str(rec.get("fecha", "")),
                substance_code(rec, sustancias_by_id),
                substance_name(rec, sustancias_by_id),
                str(rec.get("lote", "")),
                substance_code_system(rec, sustancias_by_id),
                str(rec.get("costo_unitario", "")),
                str(rec.get("costo_total", "")),
                str(rec.get("factura", "")),
                rec.get("cantidad", ""),
                str(rec.get("presentacion", "")),
                rec.get("total", ""),
                lkp.to_name("unidades", rec.get("id_unidad"), str(rec.get("unidad", ""))),
                str(rec.get("concentracion", "")),
                str(rec.get("densidad", "")),
                lkp.to_name("proveedores", rec.get("id_proveedor"), str(rec.get("fabricante", rec.get("proveedor", "")))),
                _bool_label(rec.get("certificado", False)),
                _bool_label(rec.get("msds", False)),
                str(rec.get("fecha_vencimiento", "")),
                str(rec.get("fecha_documento", "")),
                str(rec.get("vigencia_documento", "")),
                location_name(rec, locations_by_key),
                lkp.to_name("condiciones", rec.get("id_condicion_almacenamiento"), str(rec.get("condicion_almacenamiento", ""))),
                str(rec.get("observaciones", "")),
            ]
        )
    return rows


def _movement_key(record: dict) -> tuple[object, str]:
    substance_key = record.get("id_sustancia", record.get("codigo", ""))
    lote = str(record.get("lote", "")).strip()
    return substance_key, lote


def _parse_iso_date(value: str) -> date | None:
    raw = str(value or "").strip()
    if not raw:
        return None
    try:
        return datetime.strptime(raw, "%Y-%m-%d").date()
    except ValueError:
        return None


def _safe_float(value) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _compute_salida_stock_snapshots(entradas: list[dict], salidas: list[dict]) -> dict[int, tuple[float, float]]:
    events_by_key: dict[tuple[object, str], list[tuple[date, int, int, float]]] = {}

    for rec in entradas:
        d = _parse_iso_date(rec.get("fecha", ""))
        if d is None:
            continue
        key = _movement_key(rec)
        qty = _safe_float(rec.get("total", rec.get("cantidad", 0)))
        events_by_key.setdefault(key, []).append((d, 0, int(rec.get("id", 0) or 0), qty))

    for rec in salidas:
        d = _parse_iso_date(rec.get("fecha_salida", ""))
        if d is None:
            continue
        key = _movement_key(rec)
        qty = _safe_float(rec.get("cantidad", 0))
        events_by_key.setdefault(key, []).append((d, 1, int(rec.get("id", 0) or 0), qty))

    snapshots: dict[int, tuple[float, float]] = {}
    for events in events_by_key.values():
        events.sort(key=lambda item: (item[0], item[1], item[2]))
        stock = 0.0
        for _, event_type, record_id, qty in events:
            if event_type == 0:
                stock += qty
                continue
            before = stock
            after = before - qty
            snapshots[record_id] = (round(before, 6), round(after, 6))
            stock = after

    return snapshots


def _build_entrada_ref_by_key(entradas: list[dict]) -> dict[tuple[object, str], dict]:
    refs: dict[tuple[object, str], dict] = {}
    for rec in entradas:
        key = _movement_key(rec)
        refs.setdefault(key, rec)
    return refs


def _build_salidas_detailed_rows(
    salidas_periodo: list[dict],
    entradas_all: list[dict],
    salidas_all: list[dict],
    lkp: Lookups,
    sustancias_by_id: dict[int, dict],
    locations_by_key: dict[tuple[str, int], dict],
) -> list[list]:
    snapshots = _compute_salida_stock_snapshots(entradas_all, salidas_all)
    entrada_ref_by_key = _build_entrada_ref_by_key(entradas_all)

    rows: list[list] = []
    for rec in sorted(salidas_periodo, key=lambda x: (str(x.get("fecha_salida", "")), int(x.get("id", 0) or 0))):
        key = _movement_key(rec)
        entrada_ref = entrada_ref_by_key.get(key)

        fecha_salida = _parse_iso_date(rec.get("fecha_salida", ""))
        vigencia_raw = ""
        dias_vigencia = ""
        if entrada_ref is not None:
            vigencia_raw = str(entrada_ref.get("fecha_vencimiento", ""))
            fecha_vigencia = _parse_iso_date(vigencia_raw)
            if fecha_salida is not None and fecha_vigencia is not None:
                dias_vigencia = (fecha_vigencia - fecha_salida).days

        stock_before, stock_after = snapshots.get(int(rec.get("id", 0) or 0), ("", ""))

        ubicacion_origen = location_name(
            rec,
            locations_by_key,
            tipo_field="ubicacion_origen_tipo",
            id_field="id_ubicacion_origen",
            legacy_field="ubicacion_origen",
        )
        if not ubicacion_origen and entrada_ref is not None:
            ubicacion_origen = location_name(entrada_ref, locations_by_key)

        rows.append(
            [
                lkp.to_name("tipos_salida", rec.get("id_tipo_salida"), ""),
                str(rec.get("fecha_salida", "")),
                substance_code(rec, sustancias_by_id),
                substance_name(rec, sustancias_by_id),
                str(rec.get("lote", "")),
                ubicacion_origen,
                vigencia_raw,
                dias_vigencia,
                stock_before,
                stock_after,
                lkp.to_name("unidades", rec.get("id_unidad"), str(rec.get("unidad", ""))),
                rec.get("cantidad", ""),
                str(rec.get("densidad", "")),
                str(rec.get("peso_inicial", "")),
                str(rec.get("peso_final", "")),
                _bool_label(rec.get("liquido", False)),
                str(rec.get("observaciones", "")),
                _bool_label(rec.get("en_uso", False)),
            ]
        )

    return rows


def _extract_year(value: str) -> int | None:
    try:
        return datetime.strptime(str(value).strip(), "%Y-%m-%d").year
    except ValueError:
        return None


def _is_in_month(value: str, year: int, month: int) -> bool:
    try:
        d = datetime.strptime(str(value).strip(), "%Y-%m-%d").date()
        return d.year == year and d.month == month
    except ValueError:
        return False


def _month_end(year: int, month: int) -> date:
    if month == 12:
        return date(year, month, 31)
    first_next = date(year, month + 1, 1)
    return first_next - timedelta(days=1)


