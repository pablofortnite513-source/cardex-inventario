import tkinter as tk
from datetime import date
from tkinter import filedialog, messagebox, ttk

from config.config import (
    COLORS,
    ENTRADAS_FILE,
    PROVEEDORES_FILE,
    REPORTES_PATH,
    SALIDAS_FILE,
    SUSTANCIAS_FILE,
    TIPOS_ENTRADA_FILE,
    UBICACIONES_FILE,
    UBICACIONES_USO_FILE,
    UNIDADES_FILE,
)
from ui.styles import build_header
from utils.data_handler import (
    DataHandler,
    Lookups,
    build_location_indexes,
    build_substance_indexes,
    location_name,
    substance_cas,
    substance_code,
    substance_from_code,
    substance_name,
)

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


class StockAnalistaWindow:
    """Vista y exportación de stock para analista usando plantilla fija."""

    def __init__(self, parent: tk.Tk):
        self.window = tk.Toplevel(parent)
        self.window.title("Stock Analista")
        self.window.geometry("1440x620")
        self.window.configure(bg=COLORS["secondary"])

        self.search_var = tk.StringVar()
        self.tree: ttk.Treeview | None = None
        self._tree_columns: tuple[str, ...] = ()
        self._tree_base_widths: dict[str, int] = {}
        self.pagina_actual = 1
        self.total_paginas = 1
        self.por_pagina_var = tk.StringVar(value="50")
        self.pag_label: tk.Label | None = None

        self._build_ui()
        self.load_table()

    def _build_ui(self) -> None:
        wrapper = tk.Frame(self.window, bg="white", bd=1, relief="solid", padx=12, pady=12)
        wrapper.pack(expand=True, fill="both", padx=14, pady=14)

        build_header(wrapper, "Sistema de Gestión  -  Stock Analista")

        search_row = tk.Frame(wrapper, bg="white")
        search_row.pack(fill="x", pady=(6, 8))

        search_entry = tk.Entry(search_row, textvariable=self.search_var)
        search_entry.pack(side="left", fill="x", expand=True, padx=(0, 8))
        search_entry.bind("<Return>", lambda _e: self.load_table())

        tk.Button(
            search_row,
            text="Buscar",
            command=self.load_table,
            bg=COLORS["primary"],
            fg=COLORS["text_light"],
            relief="flat",
            padx=16,
            pady=5,
        ).pack(side="left", padx=(0, 8))

        tk.Button(
            search_row,
            text="Actualizar",
            command=self.load_table,
            bg=COLORS["border"],
            fg=COLORS["text_dark"],
            relief="flat",
            padx=16,
            pady=5,
        ).pack(side="left")

        columns = (
            "codigo", "cas", "nombre", "lote", "unidad", "entrada", "presentacion",
            "stock", "ubicacion", "fecha_vencimiento", "proveedor", "lote_uso",
            "tipo_entrada", "concentracion", "densidad",
        )
        self._tree_columns = columns

        tree_frame = tk.Frame(wrapper, bg="white")
        tree_frame.pack(expand=True, fill="both")
        tree_frame.rowconfigure(0, weight=1)
        tree_frame.columnconfigure(0, weight=1)

        self.tree = ttk.Treeview(tree_frame, columns=columns, show="headings", height=16)
        tree_scroll_y = ttk.Scrollbar(tree_frame, orient="vertical", command=self.tree.yview)
        tree_scroll_x = ttk.Scrollbar(tree_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=tree_scroll_y.set, xscrollcommand=tree_scroll_x.set)
        self.tree.grid(row=0, column=0, sticky="nsew")
        tree_scroll_y.grid(row=0, column=1, sticky="ns")
        tree_scroll_x.grid(row=1, column=0, sticky="ew")

        headings = {
            "codigo": "Código",
            "cas": "CAS",
            "nombre": "Nombre",
            "lote": "Lote",
            "unidad": "Unidad",
            "entrada": "Entrada",
            "presentacion": "Presentación",
            "stock": "Stock",
            "ubicacion": "Ubicación",
            "fecha_vencimiento": "F. Vencimiento",
            "proveedor": "Proveedor",
            "lote_uso": "Lote Uso",
            "tipo_entrada": "Tipo Entrada",
            "concentracion": "Concentración",
            "densidad": "Densidad",
        }
        widths = {
            "codigo": 90,
            "cas": 120,
            "nombre": 240,
            "lote": 95,
            "unidad": 70,
            "entrada": 90,
            "presentacion": 100,
            "stock": 90,
            "ubicacion": 130,
            "fecha_vencimiento": 110,
            "proveedor": 160,
            "lote_uso": 90,
            "tipo_entrada": 130,
            "concentracion": 120,
            "densidad": 90,
        }

        for col in columns:
            self.tree.heading(col, text=headings[col])
            self.tree.column(col, width=widths[col], minwidth=max(60, int(widths[col] * 0.5)), anchor="w", stretch=True)
        self._tree_base_widths = widths
        self.tree.bind("<Configure>", self._on_tree_resize, add="+")

        pag_frame = tk.Frame(wrapper, bg="white")
        pag_frame.pack(fill="x", padx=2, pady=(6, 2))
        tk.Button(
            pag_frame, text="◄ Primera", command=lambda: self._ir_pagina(1),
            bg=COLORS["border"], fg=COLORS["text_dark"], relief="flat", padx=8, pady=3,
        ).pack(side="left", padx=(0, 4))
        tk.Button(
            pag_frame, text="Anterior", command=lambda: self._ir_pagina(self.pagina_actual - 1),
            bg=COLORS["border"], fg=COLORS["text_dark"], relief="flat", padx=8, pady=3,
        ).pack(side="left", padx=(0, 4))

        self.pag_label = tk.Label(pag_frame, text="Página 1 de 1", bg="white", font=("Segoe UI", 9))
        self.pag_label.pack(side="left", padx=10)

        tk.Button(
            pag_frame, text="Siguiente", command=lambda: self._ir_pagina(self.pagina_actual + 1),
            bg=COLORS["border"], fg=COLORS["text_dark"], relief="flat", padx=8, pady=3,
        ).pack(side="left", padx=(0, 4))
        tk.Button(
            pag_frame, text="Última ►", command=lambda: self._ir_pagina(self.total_paginas),
            bg=COLORS["border"], fg=COLORS["text_dark"], relief="flat", padx=8, pady=3,
        ).pack(side="left")

        tk.Label(pag_frame, text="Mostrar:", bg="white").pack(side="left", padx=(15, 4))
        por_pagina_combo = ttk.Combobox(
            pag_frame,
            textvariable=self.por_pagina_var,
            values=["20", "50", "100", "200"],
            state="readonly",
            width=6,
        )
        por_pagina_combo.pack(side="left", padx=(0, 4))
        por_pagina_combo.bind("<<ComboboxSelected>>", lambda _e: self._cambiar_por_pagina())

        actions = tk.Frame(wrapper, bg="white")
        actions.pack(fill="x", pady=(10, 0))

        tk.Button(
            actions,
            text="Generar y guardar",
            command=self.generate_report,
            bg=COLORS["primary"],
            fg=COLORS["text_light"],
            relief="flat",
            padx=20,
            pady=7,
        ).pack(side="left")

        tk.Button(
            actions,
            text="Salir",
            command=self.window.destroy,
            bg=COLORS["border"],
            fg=COLORS["text_dark"],
            relief="flat",
            padx=20,
            pady=7,
        ).pack(side="right")

    def _on_tree_resize(self, event: tk.Event) -> None:
        if self.tree is None or not self._tree_columns:
            return
        total_base = sum(self._tree_base_widths.get(col, 1) for col in self._tree_columns)
        if total_base <= 0:
            return
        width = max(event.width - 20, 800)
        for col in self._tree_columns:
            ratio = self._tree_base_widths.get(col, 1) / total_base
            target = int(width * ratio)
            self.tree.column(col, width=max(60, target), stretch=True)

    def load_table(self) -> None:
        if self.tree is None:
            return
        query = self.search_var.get().strip().lower()
        rows = _build_stock_analista_rows()

        filtered_rows: list[list] = []
        for row in rows:
            if query and query not in str(row).lower():
                continue
            filtered_rows.append(row)

        try:
            por_pagina = max(1, int(self.por_pagina_var.get().strip()))
        except ValueError:
            por_pagina = 50
            self.por_pagina_var.set("50")

        total = len(filtered_rows)
        self.total_paginas = max(1, (total + por_pagina - 1) // por_pagina)
        self.pagina_actual = max(1, min(self.pagina_actual, self.total_paginas))
        start = (self.pagina_actual - 1) * por_pagina
        end = start + por_pagina

        self.tree.delete(*self.tree.get_children())
        for row in filtered_rows[start:end]:
            self.tree.insert("", tk.END, values=row)

        if self.pag_label is not None:
            self.pag_label.config(text=f"Página {self.pagina_actual} de {self.total_paginas}")

    def _ir_pagina(self, pagina: int) -> None:
        if pagina < 1 or pagina > self.total_paginas:
            return
        self.pagina_actual = pagina
        self.load_table()

    def _cambiar_por_pagina(self) -> None:
        self.pagina_actual = 1
        self.load_table()

    def generate_report(self) -> None:
        if load_workbook is None:
            messagebox.showerror(
                "Dependencia faltante",
                "No se encontró openpyxl. Instala con: pip install openpyxl",
                parent=self.window,
            )
            return

        template_path = REPORTES_PATH / "templates" / "template_reporteanalista.xlsx"
        if not template_path.exists():
            messagebox.showerror(
                "Plantilla no encontrada",
                f"No existe la plantilla:\n{template_path}",
                parent=self.window,
            )
            return

        default_name = f"stock_analista_{date.today().strftime('%Y%m%d')}.xlsx"
        save_path = filedialog.asksaveasfilename(
            title="Guardar reporte Stock Analista",
            initialfile=default_name,
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialdir=str(REPORTES_PATH),
            parent=self.window,
        )
        if not save_path:
            return

        try:
            data_rows = _build_stock_analista_rows()
            _render_stock_analista_sheet(template_path, save_path, data_rows)
            messagebox.showinfo(
                "Reporte generado",
                f"Excel generado correctamente:\n{save_path}",
                parent=self.window,
            )
        except Exception as e:
            messagebox.showerror(
                "Error al guardar",
                f"No se pudo generar el reporte:\n{e}",
                parent=self.window,
            )


def _safe_float(value) -> float:
    try:
        return float(value)
    except (TypeError, ValueError):
        return 0.0


def _build_stock_analista_rows() -> list[list]:
    entradas = DataHandler.get_all(ENTRADAS_FILE, "entradas")
    salidas = DataHandler.get_all(SALIDAS_FILE, "salidas")

    unidades_cat = DataHandler.load_json(UNIDADES_FILE).get("maestrasUnidades", [])
    proveedores_cat = DataHandler.load_json(PROVEEDORES_FILE).get("maestrasProveedores", [])
    sustancias_cat = DataHandler.load_json(SUSTANCIAS_FILE).get("maestrasSustancias", [])
    ubicaciones_cat = DataHandler.load_json(UBICACIONES_FILE).get("maestrasUbicaciones", [])
    ubicaciones_uso_cat = DataHandler.load_json(UBICACIONES_USO_FILE).get("maestrasUbicacionesUso", [])
    tipos_entrada_cat = DataHandler.load_json(TIPOS_ENTRADA_FILE).get("maestrasTiposEntrada", [])

    lkp = Lookups(unidades=unidades_cat, proveedores=proveedores_cat, tipos_entrada=tipos_entrada_cat)
    sustancias_by_id, sustancias_by_code = build_substance_indexes(sustancias_cat)
    locations_by_key, _ = build_location_indexes(ubicaciones_cat, ubicaciones_uso_cat)

    by_key: dict[tuple[object, str], dict] = {}

    for rec in entradas:
        if rec.get("anulado", False):
            continue

        substance_key = rec.get("id_sustancia", rec.get("codigo", ""))
        lote = str(rec.get("lote", "")).strip()
        key = (substance_key, lote)

        total_in = _safe_float(rec.get("total", rec.get("cantidad", 0)))

        if key not in by_key:
            by_key[key] = {
                "record": rec,
                "entrada": 0.0,
                "salida": 0.0,
                "en_uso": False,
            }

        by_key[key]["entrada"] += total_in

    for rec in salidas:
        if rec.get("anulado", False):
            continue

        substance_key = rec.get("id_sustancia", rec.get("codigo", ""))
        lote = str(rec.get("lote", "")).strip()
        key = (substance_key, lote)
        qty_out = _safe_float(rec.get("cantidad", 0))

        if key not in by_key:
            by_key[key] = {
                "record": rec,
                "entrada": 0.0,
                "salida": 0.0,
                "en_uso": False,
            }

        by_key[key]["salida"] += qty_out
        if qty_out > 0:
            by_key[key]["en_uso"] = True

    rows: list[list] = []
    for (_, lote), data in sorted(by_key.items(), key=lambda item: (str(item[0][0]), item[0][1])):
        record = data["record"]
        stock = round(data["entrada"] - data["salida"], 6)
        tipo_entrada = lkp.to_name("tipos_entrada", record.get("id_tipo_entrada"), "")
        lote_uso = "EN USO" if data["en_uso"] else ""

        cas = substance_cas(record, sustancias_by_id)
        if not cas:
            fallback = substance_from_code(sustancias_by_code, str(record.get("codigo", "")))
            if fallback is not None:
                cas = str(fallback.get("codigo_cas", ""))
        if not cas:
            cas = "N/D"

        row = [
            substance_code(record, sustancias_by_id),
            cas,
            substance_name(record, sustancias_by_id),
            lote,
            lkp.to_name("unidades", record.get("id_unidad"), str(record.get("unidad", ""))),
            round(data["entrada"], 6),
            str(record.get("presentacion", "")),
            stock,
            location_name(record, locations_by_key),
            str(record.get("fecha_vencimiento", "")),
            lkp.to_name("proveedores", record.get("id_proveedor"), str(record.get("fabricante", record.get("proveedor", "")))),
            lote_uso,
            tipo_entrada,
            str(record.get("concentracion", "")),
            str(record.get("densidad", "")),
        ]
        rows.append(row)

    return rows


def _render_stock_analista_sheet(template_path, save_path, data_rows: list[list]) -> None:
    template_wb = load_workbook(template_path)
    ws = template_wb.active

    template_row = 2
    start_row = 2
    max_cols = 15

    style_templates = []
    for col in range(1, max_cols + 1):
        source = ws.cell(row=template_row, column=col)
        style_templates.append(
            {
                "font": source.font.copy() if source.font else None,
                "border": source.border.copy() if source.border else None,
                "fill": source.fill.copy() if source.fill else None,
                "number_format": source.number_format,
                "alignment": source.alignment.copy() if source.alignment else None,
            }
        )

    for i, row_data in enumerate(data_rows):
        current_row = start_row + i
        ws.insert_rows(current_row)

        for col, value in enumerate(row_data, start=1):
            target = ws.cell(row=current_row, column=col, value=value)
            if col <= len(style_templates):
                st = style_templates[col - 1]
                target.font = st["font"]
                target.border = st["border"]
                target.fill = st["fill"]
                target.number_format = st["number_format"]
                target.alignment = st["alignment"]

    if data_rows:
        ws.delete_rows(template_row + len(data_rows))

    template_wb.save(save_path)
